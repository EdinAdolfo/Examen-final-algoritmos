import openpyxl
import os
from openpyxl.utils import get_column_letter

def update_excel(file_path):
    # Verificar si el archivo existe
    if not os.path.exists(file_path):
        print("El archivo no existe. Creando un nuevo archivo.")
        wb = openpyxl.Workbook()
        sheet = wb.active
        # Agregar encabezado
        sheet.append(["Nombre", "Nota", "Asignatura"])
        wb.save(file_path)
    else:
        print("El archivo ya existe. Cargando el archivo.")
        wb = openpyxl.load_workbook(file_path)

    # Cargar hoja activa
    sheet = wb.active

    # Agregar nueva fila con datos
    new_row = ["Juan", "80", "Matemáticas"]
    sheet.append(new_row)

    # Ajustar el tamaño de las columnas según el contenido
    for col in sheet.columns:
        max_length = 0
        column = col[0].column
        for cell in col:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = max_length + 2
        sheet.column_dimensions[get_column_letter(column)].width = adjusted_width

    # Guardar el archivo y manejar posibles errores
    try:
        wb.save(file_path)
        print("Los datos fueron agregados correctamente al archivo Excel.")
    except PermissionError:
        print("Error: No se puede guardar el archivo. Asegúrate de que esté cerrado.")
    
    # Cerrar el archivo
    wb.close()

# Llamar a la función con el nombre del archivo
file_path = "notas.xlsx"
update_excel(file_path)
